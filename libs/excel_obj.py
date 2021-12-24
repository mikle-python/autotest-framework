from openpyxl import load_workbook


class ExcelObj(object):
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = load_workbook(file_name)

    @property
    def sheet_names(self):
        return self.workbook.sheetnames

    def get_sheet_nrows(self, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        return sheet_obj.max_row

    def get_sheet_ncols(self, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        return sheet_obj.max_column

    def get_cell_value(self, rowx, colx, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        return sheet_obj.cell(rowx, colx).value

    def get_rows_values(self, rowx, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        return [cell.value for cell in [row for row in sheet_obj.rows][rowx]]

    def get_cols_values(self, colx, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        return [cell.value for cell in [col for col in sheet_obj.columns][colx]]

    def get_cell_row(self, value, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        i = 1
        for row in sheet_obj.rows:
            for cell in row:
                if value == cell.value:
                    return i
            i = i + 1

    def get_cell_col(self, value, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        i = 1
        for col in sheet_obj.columns:
            for cell in col:
                if value == cell.value:
                    return i
            i = i + 1

    def write_cell(self, rowx, colx, value, sheet_name):
        sheet_obj = self.workbook[sheet_name]
        sheet_obj.cell(rowx, colx, value)
        self.workbook.save(self.file_name)


if __name__ == '__main__':
    excel_obj = ExcelObj('C:/test/newben_api_amd.xlsx')
    sheet = 'User'
    key_list = excel_obj.get_rows_values(0, sheet)
    for nrow in range(1, excel_obj.get_sheet_nrows(sheet)):
        print(excel_obj.get_rows_values(nrow, sheet))